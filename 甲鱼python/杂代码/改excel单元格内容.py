"""使用openpyxl保存后会导致excel里的图片消失"""

import openpyxl

# 文件路径，python中路径要用\\
d = r'D:\杂项文件\手机3.0版修改_B优先级.xlsx'

# 打开工作薄
wb = openpyxl.load_workbook(d)

# 打开sheet工作表
ws = wb.worksheets[1]

# 获取某一个单元格值
#print(ws["B10"].value)

# 修改单元格内容
#ws["A4"] = 123

# 保存修改
#wb.save(d)

# 读每一个单元格内容，修改后再写入
a = 11
for a in range(a,99):
    # 读这一单元格的值
    atext = ws.cell(row=a, column=2).value
    print(atext)

    # 以"."为分隔符
    btext = atext.split(".")
    print(btext[1])

    # 将"."之后的字符录入到单元格里
    ws.cell(row=a, column=2).value = btext[1]
    a += 1

wb.save(d)
